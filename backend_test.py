#!/usr/bin/env python3
"""
Comprehensive Backend API Test Suite for Moviq
Tests all endpoints at REACT_APP_BACKEND_URL/api
"""

import requests
import json
import uuid
from typing import Dict, Optional, Any

# Backend URL from frontend/.env
BASE_URL = "https://video-replica-ai.preview.emergentagent.com/api"

# Test credentials
ADMIN_EMAIL = "admin@moviq.in"
ADMIN_PASSWORD = "demo1234"
AGENCY_EMAIL = "saurav@brightads.in"
AGENCY_PASSWORD = "demo1234"

# Global tokens
admin_token: Optional[str] = None
agency_token: Optional[str] = None

# Test results tracking
test_results = {
    "passed": [],
    "failed": [],
    "total": 0
}


def log_test(name: str, passed: bool, details: str = ""):
    """Log test result"""
    test_results["total"] += 1
    status = "✅ PASS" if passed else "❌ FAIL"
    result = f"{status}: {name}"
    if details:
        result += f" - {details}"
    print(result)
    
    if passed:
        test_results["passed"].append(name)
    else:
        test_results["failed"].append({"name": name, "details": details})


def make_request(method: str, endpoint: str, token: Optional[str] = None, 
                 json_data: Optional[Dict] = None, params: Optional[Dict] = None) -> requests.Response:
    """Make HTTP request with optional auth"""
    url = f"{BASE_URL}{endpoint}"
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    
    try:
        if method == "GET":
            return requests.get(url, headers=headers, params=params, timeout=10)
        elif method == "POST":
            return requests.post(url, headers=headers, json=json_data, timeout=10)
        elif method == "DELETE":
            return requests.delete(url, headers=headers, timeout=10)
        elif method == "PATCH":
            return requests.patch(url, headers=headers, json=json_data, timeout=10)
    except Exception as e:
        print(f"Request error: {e}")
        raise


# ============================================================================
# AUTH TESTS
# ============================================================================

def test_auth_login_admin_valid():
    """Test admin login with valid credentials"""
    global admin_token
    resp = make_request("POST", "/auth/login", json_data={
        "email": ADMIN_EMAIL,
        "password": ADMIN_PASSWORD
    })
    
    if resp.status_code == 200:
        data = resp.json()
        if "token" in data and "user" in data:
            admin_token = data["token"]
            user = data["user"]
            if user["email"] == ADMIN_EMAIL and user["role"] == "admin":
                log_test("Auth: Admin login valid", True)
                return
    
    log_test("Auth: Admin login valid", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_auth_login_agency_valid():
    """Test agency login with valid credentials"""
    global agency_token
    resp = make_request("POST", "/auth/login", json_data={
        "email": AGENCY_EMAIL,
        "password": AGENCY_PASSWORD
    })
    
    if resp.status_code == 200:
        data = resp.json()
        if "token" in data and "user" in data:
            agency_token = data["token"]
            user = data["user"]
            if user["email"] == AGENCY_EMAIL and user["role"] == "agency" and user.get("agencyId") == "a1":
                log_test("Auth: Agency login valid", True)
                return
    
    log_test("Auth: Agency login valid", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_auth_login_invalid_password():
    """Test login with wrong password returns 401"""
    resp = make_request("POST", "/auth/login", json_data={
        "email": ADMIN_EMAIL,
        "password": "wrongpassword"
    })
    
    if resp.status_code == 401:
        log_test("Auth: Invalid password returns 401", True)
    else:
        log_test("Auth: Invalid password returns 401", False, f"Expected 401, got {resp.status_code}")


def test_auth_me_with_token():
    """Test /auth/me with valid token"""
    resp = make_request("GET", "/auth/me", token=admin_token)
    
    if resp.status_code == 200:
        user = resp.json()
        if user["email"] == ADMIN_EMAIL and user["role"] == "admin":
            log_test("Auth: /me with valid token", True)
            return
    
    log_test("Auth: /me with valid token", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_auth_me_without_token():
    """Test /auth/me without token returns 401"""
    resp = make_request("GET", "/auth/me")
    
    if resp.status_code == 401:
        log_test("Auth: /me without token returns 401", True)
    else:
        log_test("Auth: /me without token returns 401", False, f"Expected 401, got {resp.status_code}")


# ============================================================================
# AGENCIES TESTS
# ============================================================================

def test_agencies_get_list():
    """Test GET /agencies returns 6 seeded agencies"""
    resp = make_request("GET", "/agencies", token=admin_token)
    
    if resp.status_code == 200:
        agencies = resp.json()
        if isinstance(agencies, list) and len(agencies) == 6:
            log_test("Agencies: GET list returns 6 agencies", True)
            return
        log_test("Agencies: GET list returns 6 agencies", False, f"Expected 6 agencies, got {len(agencies)}")
    else:
        log_test("Agencies: GET list returns 6 agencies", False, f"Status: {resp.status_code}")


def test_agencies_post_as_admin():
    """Test POST /agencies as admin creates new agency"""
    resp = make_request("POST", "/agencies", token=admin_token, json_data={
        "name": "Test Agency Ltd",
        "head": "John Doe",
        "email": "john@testagency.com",
        "phone": "9876543210",
        "city": "Mumbai",
        "plan": "Enterprise"
    })
    
    if resp.status_code == 200:
        agency = resp.json()
        if agency.get("name") == "Test Agency Ltd" and "id" in agency:
            log_test("Agencies: POST as admin creates agency", True)
            return
    
    log_test("Agencies: POST as admin creates agency", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_agencies_post_as_agency_forbidden():
    """Test POST /agencies as agency user returns 403"""
    resp = make_request("POST", "/agencies", token=agency_token, json_data={
        "name": "Unauthorized Agency",
        "email": "test@test.com"
    })
    
    if resp.status_code == 403:
        log_test("Agencies: POST as agency returns 403", True)
    else:
        log_test("Agencies: POST as agency returns 403", False, f"Expected 403, got {resp.status_code}")


def test_agencies_get_by_id():
    """Test GET /agencies/a1 returns BrightAds Media"""
    resp = make_request("GET", "/agencies/a1", token=admin_token)
    
    if resp.status_code == 200:
        agency = resp.json()
        if agency.get("id") == "a1" and agency.get("name") == "BrightAds Media":
            log_test("Agencies: GET by id returns correct agency", True)
            return
    
    log_test("Agencies: GET by id returns correct agency", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


# ============================================================================
# CAMPAIGNS TESTS
# ============================================================================

def test_campaigns_get_as_admin():
    """Test GET /campaigns as admin returns all 6 campaigns"""
    resp = make_request("GET", "/campaigns", token=admin_token)
    
    if resp.status_code == 200:
        campaigns = resp.json()
        if isinstance(campaigns, list) and len(campaigns) >= 6:
            log_test("Campaigns: GET as admin returns all campaigns", True)
            return
        log_test("Campaigns: GET as admin returns all campaigns", False, f"Expected >=6 campaigns, got {len(campaigns)}")
    else:
        log_test("Campaigns: GET as admin returns all campaigns", False, f"Status: {resp.status_code}")


def test_campaigns_get_as_agency():
    """Test GET /campaigns as agency returns only agencyId=a1 campaigns"""
    resp = make_request("GET", "/campaigns", token=agency_token)
    
    if resp.status_code == 200:
        campaigns = resp.json()
        if isinstance(campaigns, list):
            # All campaigns should have agencyId=a1
            all_a1 = all(c.get("agencyId") == "a1" for c in campaigns)
            if all_a1 and len(campaigns) == 2:
                log_test("Campaigns: GET as agency returns only a1 campaigns", True)
                return
            log_test("Campaigns: GET as agency returns only a1 campaigns", False, 
                    f"Expected 2 a1 campaigns, got {len(campaigns)}, all_a1={all_a1}")
        else:
            log_test("Campaigns: GET as agency returns only a1 campaigns", False, "Response not a list")
    else:
        log_test("Campaigns: GET as agency returns only a1 campaigns", False, f"Status: {resp.status_code}")


def test_campaigns_get_filter_by_agency():
    """Test GET /campaigns?agency_id=a2 filters correctly"""
    resp = make_request("GET", "/campaigns", token=admin_token, params={"agency_id": "a2"})
    
    if resp.status_code == 200:
        campaigns = resp.json()
        if isinstance(campaigns, list):
            all_a2 = all(c.get("agencyId") == "a2" for c in campaigns)
            if all_a2:
                log_test("Campaigns: GET with agency_id filter works", True)
                return
            log_test("Campaigns: GET with agency_id filter works", False, "Not all campaigns have agencyId=a2")
        else:
            log_test("Campaigns: GET with agency_id filter works", False, "Response not a list")
    else:
        log_test("Campaigns: GET with agency_id filter works", False, f"Status: {resp.status_code}")


def test_campaigns_post_as_agency():
    """Test POST /campaigns as agency creates campaign"""
    resp = make_request("POST", "/campaigns", token=agency_token, json_data={
        "title": "Test Campaign",
        "brand": "Test Brand",
        "mediaType": "Bus Branding",
        "city": "Delhi",
        "totalTasks": 10,
        "budget": 50000,
        "startDate": "2024-01-01",
        "endDate": "2024-03-31"
    })
    
    if resp.status_code == 200:
        campaign = resp.json()
        if campaign.get("title") == "Test Campaign" and campaign.get("agencyId") == "a1":
            log_test("Campaigns: POST as agency creates campaign", True)
            return
    
    log_test("Campaigns: POST as agency creates campaign", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_campaigns_post_as_admin_forbidden():
    """Test POST /campaigns as admin returns 403"""
    resp = make_request("POST", "/campaigns", token=admin_token, json_data={
        "title": "Admin Campaign",
        "brand": "Test",
        "mediaType": "Billboard",
        "city": "Mumbai"
    })
    
    if resp.status_code == 403:
        log_test("Campaigns: POST as admin returns 403", True)
    else:
        log_test("Campaigns: POST as admin returns 403", False, f"Expected 403, got {resp.status_code}")


# ============================================================================
# TASKS TESTS
# ============================================================================

def test_tasks_get_as_admin():
    """Test GET /tasks as admin returns all 40 tasks"""
    resp = make_request("GET", "/tasks", token=admin_token)
    
    if resp.status_code == 200:
        tasks = resp.json()
        if isinstance(tasks, list) and len(tasks) == 40:
            log_test("Tasks: GET as admin returns 40 tasks", True)
            return
        log_test("Tasks: GET as admin returns 40 tasks", False, f"Expected 40 tasks, got {len(tasks)}")
    else:
        log_test("Tasks: GET as admin returns 40 tasks", False, f"Status: {resp.status_code}")


def test_tasks_get_as_agency():
    """Test GET /tasks as agency returns only agencyId=a1 tasks"""
    resp = make_request("GET", "/tasks", token=agency_token)
    
    if resp.status_code == 200:
        tasks = resp.json()
        if isinstance(tasks, list):
            all_a1 = all(t.get("agencyId") == "a1" for t in tasks)
            if all_a1:
                log_test("Tasks: GET as agency returns only a1 tasks", True)
                return
            log_test("Tasks: GET as agency returns only a1 tasks", False, "Not all tasks have agencyId=a1")
        else:
            log_test("Tasks: GET as agency returns only a1 tasks", False, "Response not a list")
    else:
        log_test("Tasks: GET as agency returns only a1 tasks", False, f"Status: {resp.status_code}")


def test_tasks_filter_by_status():
    """Test GET /tasks?status_=flagged filters correctly"""
    resp = make_request("GET", "/tasks", token=admin_token, params={"status_": "flagged"})
    
    if resp.status_code == 200:
        tasks = resp.json()
        if isinstance(tasks, list):
            all_flagged = all(t.get("status") == "flagged" for t in tasks)
            if all_flagged:
                log_test("Tasks: Filter by status works", True)
                return
            log_test("Tasks: Filter by status works", False, "Not all tasks have status=flagged")
        else:
            log_test("Tasks: Filter by status works", False, "Response not a list")
    else:
        log_test("Tasks: Filter by status works", False, f"Status: {resp.status_code}")


def test_tasks_filter_by_city():
    """Test GET /tasks?city=Bengaluru filters correctly"""
    resp = make_request("GET", "/tasks", token=admin_token, params={"city": "Bengaluru"})
    
    if resp.status_code == 200:
        tasks = resp.json()
        if isinstance(tasks, list):
            all_bengaluru = all(t.get("city") == "Bengaluru" for t in tasks)
            if all_bengaluru:
                log_test("Tasks: Filter by city works", True)
                return
            log_test("Tasks: Filter by city works", False, "Not all tasks have city=Bengaluru")
        else:
            log_test("Tasks: Filter by city works", False, "Response not a list")
    else:
        log_test("Tasks: Filter by city works", False, f"Status: {resp.status_code}")


def test_task_patch_approve_as_agency():
    """Test PATCH /tasks/{id} with status=approved as agency user"""
    # Get a task belonging to agency a1
    resp = make_request("GET", "/tasks", token=agency_token)
    if resp.status_code != 200:
        log_test("Tasks: PATCH approve as agency", False, "Failed to get tasks")
        return
    
    tasks = resp.json()
    if not tasks:
        log_test("Tasks: PATCH approve as agency", False, "No tasks found for agency")
        return
    
    task_id = tasks[0]["id"]
    
    # PATCH to approve
    resp = make_request("PATCH", f"/tasks/{task_id}", token=agency_token, 
                       json_data={"status": "approved"})
    
    if resp.status_code == 200:
        task = resp.json()
        if task.get("status") == "approved" and task.get("flagReason") is None:
            log_test("Tasks: PATCH approve as agency", True)
            return
        log_test("Tasks: PATCH approve as agency", False, 
                f"Status={task.get('status')}, flagReason={task.get('flagReason')}")
    else:
        log_test("Tasks: PATCH approve as agency", False, 
                f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_task_patch_flag_with_reason_as_agency():
    """Test PATCH /tasks/{id} with status=flagged and flagReason as agency user"""
    # Get a task belonging to agency a1
    resp = make_request("GET", "/tasks", token=agency_token)
    if resp.status_code != 200:
        log_test("Tasks: PATCH flag with reason as agency", False, "Failed to get tasks")
        return
    
    tasks = resp.json()
    if not tasks:
        log_test("Tasks: PATCH flag with reason as agency", False, "No tasks found for agency")
        return
    
    task_id = tasks[0]["id"]
    
    # PATCH to flag with reason
    resp = make_request("PATCH", f"/tasks/{task_id}", token=agency_token, 
                       json_data={"status": "flagged", "flagReason": "Re-shoot requested by agency"})
    
    if resp.status_code == 200:
        task = resp.json()
        if task.get("status") == "flagged" and task.get("flagReason") == "Re-shoot requested by agency":
            log_test("Tasks: PATCH flag with reason as agency", True)
            return
        log_test("Tasks: PATCH flag with reason as agency", False, 
                f"Status={task.get('status')}, flagReason={task.get('flagReason')}")
    else:
        log_test("Tasks: PATCH flag with reason as agency", False, 
                f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_task_patch_empty_body():
    """Test PATCH /tasks/{id} with empty body returns 400"""
    # Get a task belonging to agency a1
    resp = make_request("GET", "/tasks", token=agency_token)
    if resp.status_code != 200:
        log_test("Tasks: PATCH empty body returns 400", False, "Failed to get tasks")
        return
    
    tasks = resp.json()
    if not tasks:
        log_test("Tasks: PATCH empty body returns 400", False, "No tasks found for agency")
        return
    
    task_id = tasks[0]["id"]
    
    # PATCH with empty body
    resp = make_request("PATCH", f"/tasks/{task_id}", token=agency_token, json_data={})
    
    if resp.status_code == 400:
        log_test("Tasks: PATCH empty body returns 400", True)
    else:
        log_test("Tasks: PATCH empty body returns 400", False, 
                f"Expected 400, got {resp.status_code}")


def test_task_patch_nonexistent_id():
    """Test PATCH /tasks/nonexistent returns 404"""
    resp = make_request("PATCH", "/tasks/nonexistent_task_id_12345", token=agency_token, 
                       json_data={"status": "approved"})
    
    if resp.status_code == 404:
        log_test("Tasks: PATCH nonexistent id returns 404", True)
    else:
        log_test("Tasks: PATCH nonexistent id returns 404", False, 
                f"Expected 404, got {resp.status_code}")


def test_task_patch_other_agency_task_forbidden():
    """Test PATCH /tasks/{id} for another agency's task returns 403"""
    # Get all tasks as admin to find a task NOT belonging to a1
    resp = make_request("GET", "/tasks", token=admin_token)
    if resp.status_code != 200:
        log_test("Tasks: PATCH other agency task returns 403", False, "Failed to get tasks as admin")
        return
    
    tasks = resp.json()
    other_agency_task = None
    for task in tasks:
        if task.get("agencyId") != "a1":
            other_agency_task = task
            break
    
    if not other_agency_task:
        log_test("Tasks: PATCH other agency task returns 403", False, "No tasks found for other agencies")
        return
    
    # Try to PATCH as agency a1 user
    resp = make_request("PATCH", f"/tasks/{other_agency_task['id']}", token=agency_token, 
                       json_data={"status": "approved"})
    
    if resp.status_code == 403:
        log_test("Tasks: PATCH other agency task returns 403", True)
    else:
        log_test("Tasks: PATCH other agency task returns 403", False, 
                f"Expected 403, got {resp.status_code}")


def test_task_patch_as_admin():
    """Test PATCH /tasks/{id} as admin can update any task"""
    # Get any task
    resp = make_request("GET", "/tasks", token=admin_token)
    if resp.status_code != 200:
        log_test("Tasks: PATCH as admin", False, "Failed to get tasks")
        return
    
    tasks = resp.json()
    if not tasks:
        log_test("Tasks: PATCH as admin", False, "No tasks found")
        return
    
    task_id = tasks[0]["id"]
    
    # PATCH as admin
    resp = make_request("PATCH", f"/tasks/{task_id}", token=admin_token, 
                       json_data={"status": "approved"})
    
    if resp.status_code == 200:
        task = resp.json()
        if task.get("status") == "approved":
            log_test("Tasks: PATCH as admin", True)
            return
        log_test("Tasks: PATCH as admin", False, f"Status not updated: {task.get('status')}")
    else:
        log_test("Tasks: PATCH as admin", False, 
                f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_task_patch_without_token():
    """Test PATCH /tasks/{id} without token returns 401"""
    resp = make_request("PATCH", "/tasks/t1", token=None, json_data={"status": "approved"})
    
    if resp.status_code == 401:
        log_test("Tasks: PATCH without token returns 401", True)
    else:
        log_test("Tasks: PATCH without token returns 401", False, 
                f"Expected 401, got {resp.status_code}")


def test_task_patch_persistence():
    """Test PATCH /tasks/{id} persists changes"""
    # Get a task belonging to agency a1
    resp = make_request("GET", "/tasks", token=agency_token)
    if resp.status_code != 200:
        log_test("Tasks: PATCH persistence check", False, "Failed to get tasks")
        return
    
    tasks = resp.json()
    if not tasks:
        log_test("Tasks: PATCH persistence check", False, "No tasks found for agency")
        return
    
    task_id = tasks[0]["id"]
    
    # PATCH to approved
    resp = make_request("PATCH", f"/tasks/{task_id}", token=agency_token, 
                       json_data={"status": "approved"})
    
    if resp.status_code != 200:
        log_test("Tasks: PATCH persistence check", False, "PATCH failed")
        return
    
    # GET tasks again and verify the change persisted
    resp = make_request("GET", "/tasks", token=agency_token)
    if resp.status_code == 200:
        tasks = resp.json()
        updated_task = next((t for t in tasks if t["id"] == task_id), None)
        if updated_task and updated_task.get("status") == "approved":
            log_test("Tasks: PATCH persistence check", True)
            return
        log_test("Tasks: PATCH persistence check", False, 
                f"Status not persisted: {updated_task.get('status') if updated_task else 'task not found'}")
    else:
        log_test("Tasks: PATCH persistence check", False, "Failed to verify persistence")


# ============================================================================
# USERS TESTS
# ============================================================================

def test_users_get_field():
    """Test GET /users?role=field returns 6 field executives"""
    resp = make_request("GET", "/users", token=admin_token, params={"role": "field"})
    
    if resp.status_code == 200:
        users = resp.json()
        if isinstance(users, list) and len(users) == 6:
            log_test("Users: GET role=field returns 6 users", True)
            return
        log_test("Users: GET role=field returns 6 users", False, f"Expected 6, got {len(users)}")
    else:
        log_test("Users: GET role=field returns 6 users", False, f"Status: {resp.status_code}")


def test_users_get_supervisor():
    """Test GET /users?role=supervisor returns 3 supervisors"""
    resp = make_request("GET", "/users", token=admin_token, params={"role": "supervisor"})
    
    if resp.status_code == 200:
        users = resp.json()
        if isinstance(users, list) and len(users) == 3:
            log_test("Users: GET role=supervisor returns 3 users", True)
            return
        log_test("Users: GET role=supervisor returns 3 users", False, f"Expected 3, got {len(users)}")
    else:
        log_test("Users: GET role=supervisor returns 3 users", False, f"Status: {resp.status_code}")


def test_users_get_admin():
    """Test GET /users?role=admin returns 3 hardcoded admins"""
    resp = make_request("GET", "/users", token=admin_token, params={"role": "admin"})
    
    if resp.status_code == 200:
        users = resp.json()
        if isinstance(users, list) and len(users) == 3:
            log_test("Users: GET role=admin returns 3 users", True)
            return
        log_test("Users: GET role=admin returns 3 users", False, f"Expected 3, got {len(users)}")
    else:
        log_test("Users: GET role=admin returns 3 users", False, f"Status: {resp.status_code}")


def test_users_get_agency():
    """Test GET /users?role=agency returns derived from agencies"""
    resp = make_request("GET", "/users", token=admin_token, params={"role": "agency"})
    
    if resp.status_code == 200:
        users = resp.json()
        if isinstance(users, list) and len(users) >= 6:
            log_test("Users: GET role=agency returns agency users", True)
            return
        log_test("Users: GET role=agency returns agency users", False, f"Expected >=6, got {len(users)}")
    else:
        log_test("Users: GET role=agency returns agency users", False, f"Status: {resp.status_code}")


# ============================================================================
# FRAUD ALERTS TESTS
# ============================================================================

def test_fraud_alerts_get():
    """Test GET /fraud-alerts returns 5 alerts"""
    resp = make_request("GET", "/fraud-alerts", token=admin_token)
    
    if resp.status_code == 200:
        alerts = resp.json()
        if isinstance(alerts, list) and len(alerts) >= 5:
            log_test("Fraud Alerts: GET returns alerts", True)
            return
        log_test("Fraud Alerts: GET returns alerts", False, f"Expected >=5, got {len(alerts)}")
    else:
        log_test("Fraud Alerts: GET returns alerts", False, f"Status: {resp.status_code}")


def test_fraud_alerts_resolve_existing():
    """Test POST /fraud-alerts/f1/resolve returns ok"""
    resp = make_request("POST", "/fraud-alerts/f1/resolve", token=admin_token)
    
    if resp.status_code == 200:
        data = resp.json()
        if data.get("ok") is True:
            # Verify it's removed
            resp2 = make_request("GET", "/fraud-alerts", token=admin_token)
            if resp2.status_code == 200:
                alerts = resp2.json()
                f1_exists = any(a.get("id") == "f1" for a in alerts)
                if not f1_exists:
                    log_test("Fraud Alerts: Resolve removes alert", True)
                    return
                log_test("Fraud Alerts: Resolve removes alert", False, "Alert f1 still exists after resolve")
                return
    
    log_test("Fraud Alerts: Resolve removes alert", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_fraud_alerts_resolve_nonexistent():
    """Test POST /fraud-alerts/nonexistent/resolve returns 404"""
    resp = make_request("POST", "/fraud-alerts/nonexistent123/resolve", token=admin_token)
    
    if resp.status_code == 404:
        log_test("Fraud Alerts: Resolve nonexistent returns 404", True)
    else:
        log_test("Fraud Alerts: Resolve nonexistent returns 404", False, f"Expected 404, got {resp.status_code}")


# ============================================================================
# MEDIA TYPES TESTS
# ============================================================================

def test_media_types_get():
    """Test GET /media-types returns 12 seeded types"""
    resp = make_request("GET", "/media-types", token=admin_token)
    
    if resp.status_code == 200:
        types = resp.json()
        if isinstance(types, list) and len(types) >= 12:
            log_test("Media Types: GET returns types", True)
            return
        log_test("Media Types: GET returns types", False, f"Expected >=12, got {len(types)}")
    else:
        log_test("Media Types: GET returns types", False, f"Status: {resp.status_code}")


def test_media_types_post_as_admin():
    """Test POST /media-types as admin creates new type"""
    resp = make_request("POST", "/media-types", token=admin_token, json_data={
        "label": "Airport Advertising",
        "category": "Outdoor"
    })
    
    if resp.status_code == 200:
        media_type = resp.json()
        if media_type.get("label") == "Airport Advertising" and "key" in media_type:
            log_test("Media Types: POST as admin creates type", True)
            return
    
    log_test("Media Types: POST as admin creates type", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_media_types_post_as_agency_forbidden():
    """Test POST /media-types as agency returns 403"""
    resp = make_request("POST", "/media-types", token=agency_token, json_data={
        "label": "Unauthorized Type",
        "category": "Test"
    })
    
    if resp.status_code == 403:
        log_test("Media Types: POST as agency returns 403", True)
    else:
        log_test("Media Types: POST as agency returns 403", False, f"Expected 403, got {resp.status_code}")


def test_media_types_post_duplicate():
    """Test POST duplicate media type returns 409"""
    # Try to create the same type again
    resp = make_request("POST", "/media-types", token=admin_token, json_data={
        "label": "Airport Advertising",
        "category": "Outdoor"
    })
    
    if resp.status_code == 409:
        log_test("Media Types: POST duplicate returns 409", True)
    else:
        log_test("Media Types: POST duplicate returns 409", False, f"Expected 409, got {resp.status_code}")


def test_media_types_delete_as_admin():
    """Test DELETE /media-types/{key} as admin"""
    # First get the key we just created
    resp = make_request("GET", "/media-types", token=admin_token)
    if resp.status_code == 200:
        types = resp.json()
        airport_type = next((t for t in types if t.get("label") == "Airport Advertising"), None)
        if airport_type:
            key = airport_type["key"]
            resp2 = make_request("DELETE", f"/media-types/{key}", token=admin_token)
            if resp2.status_code == 200:
                log_test("Media Types: DELETE as admin works", True)
                return
    
    log_test("Media Types: DELETE as admin works", False, "Could not delete media type")


def test_media_types_delete_as_agency_forbidden():
    """Test DELETE /media-types/{key} as agency returns 403"""
    resp = make_request("DELETE", "/media-types/bus_branding", token=agency_token)
    
    if resp.status_code == 403:
        log_test("Media Types: DELETE as agency returns 403", True)
    else:
        log_test("Media Types: DELETE as agency returns 403", False, f"Expected 403, got {resp.status_code}")


def test_media_types_delete_nonexistent():
    """Test DELETE /media-types/nonexistent returns 404"""
    resp = make_request("DELETE", "/media-types/nonexistent_key_123", token=admin_token)
    
    if resp.status_code == 404:
        log_test("Media Types: DELETE nonexistent returns 404", True)
    else:
        log_test("Media Types: DELETE nonexistent returns 404", False, f"Expected 404, got {resp.status_code}")


# ============================================================================
# VEHICLE SUBMISSIONS TESTS
# ============================================================================

def test_vehicle_submissions_post_without_auth():
    """Test POST /vehicle-submissions without auth (public endpoint)"""
    resp = make_request("POST", "/vehicle-submissions", json_data={
        "vehicle": "KA01AB1234",
        "driver_name": "Rajesh Kumar",
        "driver_phone": "9876543210",
        "photos": [{"url": "https://example.com/photo1.jpg"}],
        "gps": {"lat": 12.9716, "lng": 77.5946}
    })
    
    if resp.status_code == 200:
        submission = resp.json()
        if submission.get("vehicle") == "KA01AB1234" and "id" in submission:
            log_test("Vehicle Submissions: POST without auth works", True)
            return
    
    log_test("Vehicle Submissions: POST without auth works", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_vehicle_submissions_get_without_auth():
    """Test GET /vehicle-submissions without auth returns 401"""
    resp = make_request("GET", "/vehicle-submissions")
    
    if resp.status_code == 401:
        log_test("Vehicle Submissions: GET without auth returns 401", True)
    else:
        log_test("Vehicle Submissions: GET without auth returns 401", False, f"Expected 401, got {resp.status_code}")


def test_vehicle_submissions_get_with_auth():
    """Test GET /vehicle-submissions with auth returns list"""
    resp = make_request("GET", "/vehicle-submissions", token=admin_token)
    
    if resp.status_code == 200:
        submissions = resp.json()
        if isinstance(submissions, list):
            # Should include the one we just posted
            has_test_vehicle = any(s.get("vehicle") == "KA01AB1234" for s in submissions)
            if has_test_vehicle:
                log_test("Vehicle Submissions: GET with auth returns list", True)
                return
            log_test("Vehicle Submissions: GET with auth returns list", False, "Test vehicle not found in list")
        else:
            log_test("Vehicle Submissions: GET with auth returns list", False, "Response not a list")
    else:
        log_test("Vehicle Submissions: GET with auth returns list", False, f"Status: {resp.status_code}")


# ============================================================================
# ANALYTICS TESTS
# ============================================================================

def test_analytics_overview_as_admin():
    """Test GET /analytics/overview as admin"""
    resp = make_request("GET", "/analytics/overview", token=admin_token)
    
    if resp.status_code == 200:
        data = resp.json()
        required_keys = ["monthlyStats", "cityStats", "kpis"]
        kpi_keys = ["activeAgencies", "liveCampaigns", "tasksExecuted", "totalTasks", "totalRevenue"]
        
        has_all_keys = all(k in data for k in required_keys)
        has_all_kpis = all(k in data.get("kpis", {}) for k in kpi_keys)
        
        if has_all_keys and has_all_kpis:
            monthly = data["monthlyStats"]
            cities = data["cityStats"]
            if isinstance(monthly, list) and len(monthly) == 6 and isinstance(cities, list) and len(cities) == 8:
                log_test("Analytics: Overview as admin has correct structure", True)
                return
            log_test("Analytics: Overview as admin has correct structure", False, 
                    f"monthlyStats={len(monthly)}, cityStats={len(cities)}")
        else:
            log_test("Analytics: Overview as admin has correct structure", False, "Missing required keys")
    else:
        log_test("Analytics: Overview as admin has correct structure", False, f"Status: {resp.status_code}")


def test_analytics_overview_as_agency():
    """Test GET /analytics/overview as agency (scoped to agencyId=a1)"""
    resp = make_request("GET", "/analytics/overview", token=agency_token)
    
    if resp.status_code == 200:
        data = resp.json()
        required_keys = ["monthlyStats", "cityStats", "kpis"]
        kpi_keys = ["activeAgencies", "liveCampaigns", "tasksExecuted", "totalTasks", "totalRevenue"]
        
        has_all_keys = all(k in data for k in required_keys)
        has_all_kpis = all(k in data.get("kpis", {}) for k in kpi_keys)
        
        if has_all_keys and has_all_kpis:
            log_test("Analytics: Overview as agency has correct structure", True)
            return
        log_test("Analytics: Overview as agency has correct structure", False, "Missing required keys")
    else:
        log_test("Analytics: Overview as agency has correct structure", False, f"Status: {resp.status_code}")


# ============================================================================
# NOTIFICATIONS & BRANDS TESTS
# ============================================================================

def test_notifications_get():
    """Test GET /notifications returns 4 items"""
    resp = make_request("GET", "/notifications", token=admin_token)
    
    if resp.status_code == 200:
        notifications = resp.json()
        if isinstance(notifications, list) and len(notifications) == 4:
            log_test("Notifications: GET returns 4 items", True)
            return
        log_test("Notifications: GET returns 4 items", False, f"Expected 4, got {len(notifications)}")
    else:
        log_test("Notifications: GET returns 4 items", False, f"Status: {resp.status_code}")


def test_brands_get():
    """Test GET /brands returns 6 items"""
    resp = make_request("GET", "/brands", token=admin_token)
    
    if resp.status_code == 200:
        brands = resp.json()
        if isinstance(brands, list) and len(brands) == 6:
            log_test("Brands: GET returns 6 items", True)
            return
        log_test("Brands: GET returns 6 items", False, f"Expected 6, got {len(brands)}")
    else:
        log_test("Brands: GET returns 6 items", False, f"Status: {resp.status_code}")


# ============================================================================
# NEW FEATURES TESTS (Campaign Detail, PDF/Excel Reports, Notifications with triggers)
# ============================================================================

def test_campaign_detail_as_admin():
    """Test GET /campaigns/{cid} as admin returns full campaign detail"""
    resp = make_request("GET", "/campaigns/c1", token=admin_token)
    
    if resp.status_code == 200:
        data = resp.json()
        required_keys = ["campaign", "tasks", "team", "activity", "stats"]
        if all(k in data for k in required_keys):
            campaign = data["campaign"]
            tasks = data["tasks"]
            if campaign.get("id") == "c1" and isinstance(tasks, list):
                # Verify all tasks belong to campaign c1
                all_c1 = all(t.get("campaignId") == "c1" for t in tasks)
                if all_c1:
                    log_test("Campaign Detail: GET /campaigns/c1 as admin", True)
                    return
                log_test("Campaign Detail: GET /campaigns/c1 as admin", False, "Not all tasks belong to c1")
                return
        log_test("Campaign Detail: GET /campaigns/c1 as admin", False, f"Missing keys or wrong structure")
    else:
        log_test("Campaign Detail: GET /campaigns/c1 as admin", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_campaign_detail_as_agency_own():
    """Test GET /campaigns/{cid} as agency for own campaign returns 200"""
    resp = make_request("GET", "/campaigns/c1", token=agency_token)
    
    if resp.status_code == 200:
        data = resp.json()
        required_keys = ["campaign", "tasks", "team", "activity", "stats"]
        if all(k in data for k in required_keys):
            log_test("Campaign Detail: GET /campaigns/c1 as agency (own campaign)", True)
            return
        log_test("Campaign Detail: GET /campaigns/c1 as agency (own campaign)", False, "Missing required keys")
    else:
        log_test("Campaign Detail: GET /campaigns/c1 as agency (own campaign)", False, f"Status: {resp.status_code}")


def test_campaign_detail_as_agency_other():
    """Test GET /campaigns/{cid} as agency for other agency's campaign returns 403"""
    # c3 belongs to a2 (Metro Outdoor), agency_token is for a1 (BrightAds)
    resp = make_request("GET", "/campaigns/c3", token=agency_token)
    
    if resp.status_code == 403:
        data = resp.json()
        if "Not your campaign" in data.get("detail", ""):
            log_test("Campaign Detail: GET /campaigns/c3 as agency (other's campaign) returns 403", True)
            return
        log_test("Campaign Detail: GET /campaigns/c3 as agency (other's campaign) returns 403", False, f"Wrong error message: {data.get('detail')}")
    else:
        log_test("Campaign Detail: GET /campaigns/c3 as agency (other's campaign) returns 403", False, f"Expected 403, got {resp.status_code}")


def test_campaign_detail_not_found():
    """Test GET /campaigns/nonexistent returns 404"""
    resp = make_request("GET", "/campaigns/does-not-exist", token=admin_token)
    
    if resp.status_code == 404:
        log_test("Campaign Detail: GET /campaigns/nonexistent returns 404", True)
    else:
        log_test("Campaign Detail: GET /campaigns/nonexistent returns 404", False, f"Expected 404, got {resp.status_code}")


def test_campaign_detail_without_token():
    """Test GET /campaigns/{cid} without token returns 401"""
    resp = make_request("GET", "/campaigns/c1", token=None)
    
    if resp.status_code == 401:
        log_test("Campaign Detail: GET /campaigns/c1 without token returns 401", True)
    else:
        log_test("Campaign Detail: GET /campaigns/c1 without token returns 401", False, f"Expected 401, got {resp.status_code}")


def test_pdf_report_as_admin():
    """Test GET /campaigns/{cid}/report/pdf as admin returns valid PDF"""
    resp = make_request("GET", "/campaigns/c1/report/pdf", token=admin_token)
    
    if resp.status_code == 200:
        content_type = resp.headers.get("Content-Type", "")
        content_disp = resp.headers.get("Content-Disposition", "")
        body = resp.content
        
        # Check Content-Type starts with application/pdf
        if not content_type.startswith("application/pdf"):
            log_test("PDF Report: GET /campaigns/c1/report/pdf as admin", False, f"Wrong Content-Type: {content_type}")
            return
        
        # Check Content-Disposition contains attachment and .pdf
        if "attachment" not in content_disp or ".pdf" not in content_disp:
            log_test("PDF Report: GET /campaigns/c1/report/pdf as admin", False, f"Wrong Content-Disposition: {content_disp}")
            return
        
        # Check body starts with %PDF- (PDF magic bytes)
        if not body.startswith(b"%PDF"):
            log_test("PDF Report: GET /campaigns/c1/report/pdf as admin", False, f"Body doesn't start with %PDF, starts with: {body[:10]}")
            return
        
        # Check body is non-empty
        if len(body) == 0:
            log_test("PDF Report: GET /campaigns/c1/report/pdf as admin", False, "Empty PDF body")
            return
        
        log_test("PDF Report: GET /campaigns/c1/report/pdf as admin", True)
    else:
        log_test("PDF Report: GET /campaigns/c1/report/pdf as admin", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_pdf_report_as_agency_own():
    """Test GET /campaigns/{cid}/report/pdf as agency for own campaign returns 200"""
    resp = make_request("GET", "/campaigns/c1/report/pdf", token=agency_token)
    
    if resp.status_code == 200:
        body = resp.content
        if body.startswith(b"%PDF") and len(body) > 0:
            log_test("PDF Report: GET /campaigns/c1/report/pdf as agency (own campaign)", True)
            return
        log_test("PDF Report: GET /campaigns/c1/report/pdf as agency (own campaign)", False, "Invalid PDF")
    else:
        log_test("PDF Report: GET /campaigns/c1/report/pdf as agency (own campaign)", False, f"Status: {resp.status_code}")


def test_pdf_report_as_agency_other():
    """Test GET /campaigns/{cid}/report/pdf as agency for other's campaign returns 403"""
    resp = make_request("GET", "/campaigns/c3/report/pdf", token=agency_token)
    
    if resp.status_code == 403:
        log_test("PDF Report: GET /campaigns/c3/report/pdf as agency (other's campaign) returns 403", True)
    else:
        log_test("PDF Report: GET /campaigns/c3/report/pdf as agency (other's campaign) returns 403", False, f"Expected 403, got {resp.status_code}")


def test_pdf_report_not_found():
    """Test GET /campaigns/nonexistent/report/pdf returns 404"""
    resp = make_request("GET", "/campaigns/does-not-exist/report/pdf", token=admin_token)
    
    if resp.status_code == 404:
        log_test("PDF Report: GET /campaigns/nonexistent/report/pdf returns 404", True)
    else:
        log_test("PDF Report: GET /campaigns/nonexistent/report/pdf returns 404", False, f"Expected 404, got {resp.status_code}")


def test_pdf_report_without_token():
    """Test GET /campaigns/{cid}/report/pdf without token returns 401"""
    resp = make_request("GET", "/campaigns/c1/report/pdf", token=None)
    
    if resp.status_code == 401:
        log_test("PDF Report: GET /campaigns/c1/report/pdf without token returns 401", True)
    else:
        log_test("PDF Report: GET /campaigns/c1/report/pdf without token returns 401", False, f"Expected 401, got {resp.status_code}")


def test_excel_report_as_admin():
    """Test GET /campaigns/{cid}/report/excel as admin returns valid Excel"""
    resp = make_request("GET", "/campaigns/c1/report/excel", token=admin_token)
    
    if resp.status_code == 200:
        content_type = resp.headers.get("Content-Type", "")
        content_disp = resp.headers.get("Content-Disposition", "")
        body = resp.content
        
        # Check Content-Type includes spreadsheetml.sheet
        if "spreadsheetml.sheet" not in content_type:
            log_test("Excel Report: GET /campaigns/c1/report/excel as admin", False, f"Wrong Content-Type: {content_type}")
            return
        
        # Check Content-Disposition contains attachment
        if "attachment" not in content_disp:
            log_test("Excel Report: GET /campaigns/c1/report/excel as admin", False, f"Wrong Content-Disposition: {content_disp}")
            return
        
        # Check body starts with PK (XLSX is a zip, magic bytes 50 4B 03 04)
        if not body.startswith(b"PK"):
            log_test("Excel Report: GET /campaigns/c1/report/excel as admin", False, f"Body doesn't start with PK, starts with: {body[:10]}")
            return
        
        # Check body is non-empty
        if len(body) == 0:
            log_test("Excel Report: GET /campaigns/c1/report/excel as admin", False, "Empty Excel body")
            return
        
        # Optionally verify workbook structure with openpyxl
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(BytesIO(body))
            sheet_names = wb.sheetnames
            if "Summary" not in sheet_names or "Tasks" not in sheet_names:
                log_test("Excel Report: GET /campaigns/c1/report/excel as admin", False, f"Missing sheets. Found: {sheet_names}")
                return
        except Exception as e:
            log_test("Excel Report: GET /campaigns/c1/report/excel as admin", False, f"Failed to load workbook: {e}")
            return
        
        log_test("Excel Report: GET /campaigns/c1/report/excel as admin", True)
    else:
        log_test("Excel Report: GET /campaigns/c1/report/excel as admin", False, f"Status: {resp.status_code}, Body: {resp.text[:200]}")


def test_excel_report_as_agency_other():
    """Test GET /campaigns/{cid}/report/excel as agency for other's campaign returns 403"""
    resp = make_request("GET", "/campaigns/c3/report/excel", token=agency_token)
    
    if resp.status_code == 403:
        log_test("Excel Report: GET /campaigns/c3/report/excel as agency (other's campaign) returns 403", True)
    else:
        log_test("Excel Report: GET /campaigns/c3/report/excel as agency (other's campaign) returns 403", False, f"Expected 403, got {resp.status_code}")


def test_excel_report_not_found():
    """Test GET /campaigns/nonexistent/report/excel returns 404"""
    resp = make_request("GET", "/campaigns/does-not-exist/report/excel", token=admin_token)
    
    if resp.status_code == 404:
        log_test("Excel Report: GET /campaigns/nonexistent/report/excel returns 404", True)
    else:
        log_test("Excel Report: GET /campaigns/nonexistent/report/excel returns 404", False, f"Expected 404, got {resp.status_code}")


def test_notifications_with_triggers_agency_create():
    """Test notification trigger: Agency create"""
    # Get current notification count
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on agency create", False, "Failed to get initial notifications")
        return
    
    initial_notifications = resp.json()
    initial_count = len(initial_notifications)
    
    # Create a new agency
    resp = make_request("POST", "/agencies", token=admin_token, json_data={
        "name": f"Test Agency {uuid.uuid4().hex[:6]}",
        "head": "Test Head",
        "email": f"test{uuid.uuid4().hex[:6]}@testagency.com",
        "phone": "9876543210",
        "city": "Mumbai",
        "plan": "Enterprise"
    })
    
    if resp.status_code != 200:
        log_test("Notifications: Trigger on agency create", False, f"Failed to create agency: {resp.status_code}")
        return
    
    # Get notifications again
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on agency create", False, "Failed to get notifications after agency create")
        return
    
    new_notifications = resp.json()
    
    # Check if new notification was created
    if len(new_notifications) <= initial_count:
        log_test("Notifications: Trigger on agency create", False, f"No new notification. Before: {initial_count}, After: {len(new_notifications)}")
        return
    
    # Check if the first (newest) notification is about agency onboarding
    newest = new_notifications[0]
    if newest.get("title") == "Agency onboarded" and newest.get("type") == "info":
        # Verify it has createdAt and time fields
        if "createdAt" in newest and "time" in newest:
            log_test("Notifications: Trigger on agency create", True)
            return
        log_test("Notifications: Trigger on agency create", False, "Missing createdAt or time field")
        return
    
    log_test("Notifications: Trigger on agency create", False, f"Wrong notification. Title: {newest.get('title')}, Type: {newest.get('type')}")


def test_notifications_with_triggers_campaign_create():
    """Test notification trigger: Campaign create"""
    # Get current notification count
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on campaign create", False, "Failed to get initial notifications")
        return
    
    initial_count = len(resp.json())
    
    # Create a new campaign as agency
    resp = make_request("POST", "/campaigns", token=agency_token, json_data={
        "title": f"Test Campaign {uuid.uuid4().hex[:6]}",
        "brand": "Test Brand",
        "mediaType": "Bus Branding",
        "city": "Delhi",
        "totalTasks": 10,
        "budget": 50000,
        "startDate": "2024-01-01",
        "endDate": "2024-03-31"
    })
    
    if resp.status_code != 200:
        log_test("Notifications: Trigger on campaign create", False, f"Failed to create campaign: {resp.status_code}")
        return
    
    # Get notifications again
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on campaign create", False, "Failed to get notifications after campaign create")
        return
    
    new_notifications = resp.json()
    
    # Check if new notification was created
    if len(new_notifications) <= initial_count:
        log_test("Notifications: Trigger on campaign create", False, f"No new notification. Before: {initial_count}, After: {len(new_notifications)}")
        return
    
    # Check if the first (newest) notification is about campaign launch
    newest = new_notifications[0]
    if newest.get("title") == "Campaign launched" and newest.get("type") == "success":
        if "createdAt" in newest and "time" in newest:
            log_test("Notifications: Trigger on campaign create", True)
            return
        log_test("Notifications: Trigger on campaign create", False, "Missing createdAt or time field")
        return
    
    log_test("Notifications: Trigger on campaign create", False, f"Wrong notification. Title: {newest.get('title')}, Type: {newest.get('type')}")


def test_notifications_with_triggers_fraud_resolve():
    """Test notification trigger: Fraud alert resolve"""
    # Get a fraud alert to resolve
    resp = make_request("GET", "/fraud-alerts", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on fraud resolve", False, "Failed to get fraud alerts")
        return
    
    alerts = resp.json()
    if not alerts:
        log_test("Notifications: Trigger on fraud resolve", False, "No fraud alerts available to resolve")
        return
    
    alert_id = alerts[0]["id"]
    
    # Get current notification count
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on fraud resolve", False, "Failed to get initial notifications")
        return
    
    initial_count = len(resp.json())
    
    # Resolve the fraud alert
    resp = make_request("POST", f"/fraud-alerts/{alert_id}/resolve", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on fraud resolve", False, f"Failed to resolve fraud alert: {resp.status_code}")
        return
    
    # Get notifications again
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on fraud resolve", False, "Failed to get notifications after fraud resolve")
        return
    
    new_notifications = resp.json()
    
    # Check if new notification was created
    if len(new_notifications) <= initial_count:
        log_test("Notifications: Trigger on fraud resolve", False, f"No new notification. Before: {initial_count}, After: {len(new_notifications)}")
        return
    
    # Check if the first (newest) notification is about fraud resolution
    newest = new_notifications[0]
    if newest.get("title") == "Fraud alert resolved" and newest.get("type") == "success":
        if "createdAt" in newest and "time" in newest:
            log_test("Notifications: Trigger on fraud resolve", True)
            return
        log_test("Notifications: Trigger on fraud resolve", False, "Missing createdAt or time field")
        return
    
    log_test("Notifications: Trigger on fraud resolve", False, f"Wrong notification. Title: {newest.get('title')}, Type: {newest.get('type')}")


def test_notifications_with_triggers_vehicle_submission():
    """Test notification trigger: Vehicle submission"""
    # Get current notification count
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on vehicle submission", False, "Failed to get initial notifications")
        return
    
    initial_count = len(resp.json())
    
    # Submit a vehicle (public endpoint, no auth)
    resp = make_request("POST", "/vehicle-submissions", json_data={
        "vehicle": f"KA{uuid.uuid4().hex[:2].upper()}AB{uuid.uuid4().hex[:4].upper()}",
        "driver_name": "Test Driver",
        "driver_phone": "9876543210",
        "photos": [{"url": "https://example.com/photo1.jpg"}],
        "gps": {"lat": 12.9716, "lng": 77.5946}
    })
    
    if resp.status_code != 200:
        log_test("Notifications: Trigger on vehicle submission", False, f"Failed to submit vehicle: {resp.status_code}")
        return
    
    # Get notifications again
    resp = make_request("GET", "/notifications", token=admin_token)
    if resp.status_code != 200:
        log_test("Notifications: Trigger on vehicle submission", False, "Failed to get notifications after vehicle submission")
        return
    
    new_notifications = resp.json()
    
    # Check if new notification was created
    if len(new_notifications) <= initial_count:
        log_test("Notifications: Trigger on vehicle submission", False, f"No new notification. Before: {initial_count}, After: {len(new_notifications)}")
        return
    
    # Check if the first (newest) notification is about vehicle proof
    newest = new_notifications[0]
    if newest.get("title") == "New vehicle proof" and newest.get("type") == "info":
        if "createdAt" in newest and "time" in newest:
            log_test("Notifications: Trigger on vehicle submission", True)
            return
        log_test("Notifications: Trigger on vehicle submission", False, "Missing createdAt or time field")
        return
    
    log_test("Notifications: Trigger on vehicle submission", False, f"Wrong notification. Title: {newest.get('title')}, Type: {newest.get('type')}")


def test_notifications_sorted_by_createdat():
    """Test GET /notifications returns list sorted newest first"""
    resp = make_request("GET", "/notifications", token=admin_token)
    
    if resp.status_code == 200:
        notifications = resp.json()
        if not isinstance(notifications, list) or len(notifications) == 0:
            log_test("Notifications: Sorted by createdAt (newest first)", False, "Empty or invalid response")
            return
        
        # Check all have createdAt and time fields
        for n in notifications:
            if "createdAt" not in n or "time" not in n:
                log_test("Notifications: Sorted by createdAt (newest first)", False, f"Missing createdAt or time field in notification {n.get('id')}")
                return
        
        # Check if sorted by createdAt descending (newest first)
        created_ats = [n["createdAt"] for n in notifications]
        sorted_created_ats = sorted(created_ats, reverse=True)
        
        if created_ats == sorted_created_ats:
            log_test("Notifications: Sorted by createdAt (newest first)", True)
            return
        
        log_test("Notifications: Sorted by createdAt (newest first)", False, "Notifications not sorted by createdAt descending")
    else:
        log_test("Notifications: Sorted by createdAt (newest first)", False, f"Status: {resp.status_code}")


def test_regression_login():
    """Regression: POST /api/auth/login still works"""
    resp = make_request("POST", "/auth/login", json_data={
        "email": ADMIN_EMAIL,
        "password": ADMIN_PASSWORD
    })
    
    if resp.status_code == 200:
        data = resp.json()
        if "token" in data and "user" in data:
            log_test("Regression: POST /auth/login (admin)", True)
            return
    
    log_test("Regression: POST /auth/login (admin)", False, f"Status: {resp.status_code}")


def test_regression_campaigns_list():
    """Regression: GET /api/campaigns still works"""
    resp = make_request("GET", "/campaigns", token=admin_token)
    
    if resp.status_code == 200:
        campaigns = resp.json()
        if isinstance(campaigns, list):
            log_test("Regression: GET /campaigns", True)
            return
    
    log_test("Regression: GET /campaigns", False, f"Status: {resp.status_code}")


def test_regression_task_patch():
    """Regression: PATCH /api/tasks/{id} still works"""
    # Get a task belonging to agency a1
    resp = make_request("GET", "/tasks", token=agency_token)
    if resp.status_code != 200:
        log_test("Regression: PATCH /tasks/{id}", False, "Failed to get tasks")
        return
    
    tasks = resp.json()
    if not tasks:
        log_test("Regression: PATCH /tasks/{id}", False, "No tasks found")
        return
    
    task_id = tasks[0]["id"]
    
    # PATCH to approve
    resp = make_request("PATCH", f"/tasks/{task_id}", token=agency_token, 
                       json_data={"status": "approved"})
    
    if resp.status_code == 200:
        task = resp.json()
        if task.get("status") == "approved":
            log_test("Regression: PATCH /tasks/{id}", True)
            return
    
    log_test("Regression: PATCH /tasks/{id}", False, f"Status: {resp.status_code}")


# ============================================================================
# MAIN TEST RUNNER
# ============================================================================

def run_all_tests():
    """Run all backend API tests"""
    print("=" * 80)
    print("MOVIQ BACKEND API TEST SUITE")
    print(f"Testing: {BASE_URL}")
    print("=" * 80)
    print()
    
    # Auth tests (must run first to get tokens)
    print("--- AUTH TESTS ---")
    test_auth_login_admin_valid()
    test_auth_login_agency_valid()
    test_auth_login_invalid_password()
    test_auth_me_with_token()
    test_auth_me_without_token()
    print()
    
    # Agencies tests
    print("--- AGENCIES TESTS ---")
    test_agencies_get_list()
    test_agencies_post_as_admin()
    test_agencies_post_as_agency_forbidden()
    test_agencies_get_by_id()
    print()
    
    # Campaigns tests
    print("--- CAMPAIGNS TESTS ---")
    test_campaigns_get_as_admin()
    test_campaigns_get_as_agency()
    test_campaigns_get_filter_by_agency()
    test_campaigns_post_as_agency()
    test_campaigns_post_as_admin_forbidden()
    print()
    
    # Tasks tests
    print("--- TASKS TESTS ---")
    test_tasks_get_as_admin()
    test_tasks_get_as_agency()
    test_tasks_filter_by_status()
    test_tasks_filter_by_city()
    print()
    
    # Task PATCH tests (bug fix verification)
    print("--- TASK UPDATE (PATCH) TESTS ---")
    test_task_patch_approve_as_agency()
    test_task_patch_flag_with_reason_as_agency()
    test_task_patch_empty_body()
    test_task_patch_nonexistent_id()
    test_task_patch_other_agency_task_forbidden()
    test_task_patch_as_admin()
    test_task_patch_without_token()
    test_task_patch_persistence()
    print()
    
    # Users tests
    print("--- USERS TESTS ---")
    test_users_get_field()
    test_users_get_supervisor()
    test_users_get_admin()
    test_users_get_agency()
    print()
    
    # Fraud alerts tests
    print("--- FRAUD ALERTS TESTS ---")
    test_fraud_alerts_get()
    test_fraud_alerts_resolve_existing()
    test_fraud_alerts_resolve_nonexistent()
    print()
    
    # Media types tests
    print("--- MEDIA TYPES TESTS ---")
    test_media_types_get()
    test_media_types_post_as_admin()
    test_media_types_post_as_agency_forbidden()
    test_media_types_post_duplicate()
    test_media_types_delete_as_admin()
    test_media_types_delete_as_agency_forbidden()
    test_media_types_delete_nonexistent()
    print()
    
    # Vehicle submissions tests
    print("--- VEHICLE SUBMISSIONS TESTS ---")
    test_vehicle_submissions_post_without_auth()
    test_vehicle_submissions_get_without_auth()
    test_vehicle_submissions_get_with_auth()
    print()
    
    # Analytics tests
    print("--- ANALYTICS TESTS ---")
    test_analytics_overview_as_admin()
    test_analytics_overview_as_agency()
    print()
    
    # Notifications & Brands tests
    print("--- NOTIFICATIONS & BRANDS TESTS ---")
    test_notifications_get()
    test_brands_get()
    print()
    
    # NEW FEATURES TESTS
    print("--- NEW FEATURE: CAMPAIGN DETAIL ---")
    test_campaign_detail_as_admin()
    test_campaign_detail_as_agency_own()
    test_campaign_detail_as_agency_other()
    test_campaign_detail_not_found()
    test_campaign_detail_without_token()
    print()
    
    print("--- NEW FEATURE: PDF REPORT ---")
    test_pdf_report_as_admin()
    test_pdf_report_as_agency_own()
    test_pdf_report_as_agency_other()
    test_pdf_report_not_found()
    test_pdf_report_without_token()
    print()
    
    print("--- NEW FEATURE: EXCEL REPORT ---")
    test_excel_report_as_admin()
    test_excel_report_as_agency_other()
    test_excel_report_not_found()
    print()
    
    print("--- NEW FEATURE: NOTIFICATIONS WITH TRIGGERS ---")
    test_notifications_with_triggers_agency_create()
    test_notifications_with_triggers_campaign_create()
    test_notifications_with_triggers_fraud_resolve()
    test_notifications_with_triggers_vehicle_submission()
    test_notifications_sorted_by_createdat()
    print()
    
    print("--- REGRESSION TESTS ---")
    test_regression_login()
    test_regression_campaigns_list()
    test_regression_task_patch()
    print()
    
    # Summary
    print("=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    print(f"Total tests: {test_results['total']}")
    print(f"Passed: {len(test_results['passed'])}")
    print(f"Failed: {len(test_results['failed'])}")
    print()
    
    if test_results['failed']:
        print("FAILED TESTS:")
        for fail in test_results['failed']:
            print(f"  ❌ {fail['name']}")
            if fail['details']:
                print(f"     {fail['details']}")
        print()
    
    success_rate = (len(test_results['passed']) / test_results['total'] * 100) if test_results['total'] > 0 else 0
    print(f"Success rate: {success_rate:.1f}%")
    print("=" * 80)


if __name__ == "__main__":
    run_all_tests()
