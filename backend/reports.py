"""PDF and Excel report builders — extracted from server.py."""
import io
from datetime import datetime, timezone
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def build_pdf(campaign: Dict[str, Any], tasks: List[Dict[str, Any]]) -> bytes:
    """Return a branded PDF campaign report as bytes."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.4 * cm, rightMargin=1.4 * cm,
        topMargin=1.4 * cm, bottomMargin=1.4 * cm,
    )
    styles = getSampleStyleSheet()
    brand_color = colors.HexColor("#DC2626")
    story = []

    styles.add(ParagraphStyle(
        name="Brand", parent=styles["Normal"],
        fontSize=10, textColor=brand_color, fontName="Helvetica-Bold",
    ))
    story.append(Paragraph("MOVIQ &nbsp;·&nbsp; Intelligence in Motion", styles["Brand"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "<b>Campaign Report</b>",
        ParagraphStyle(name="H1", parent=styles["Title"], fontSize=20,
                       textColor=colors.HexColor("#0f172a")),
    ))
    story.append(Paragraph(
        campaign.get("title", ""),
        ParagraphStyle(name="Sub", parent=styles["Normal"], fontSize=13,
                       textColor=colors.HexColor("#334155")),
    ))
    story.append(Spacer(1, 14))

    meta_data = [
        ["Brand",      campaign.get("brand", "-"),     "Agency",  campaign.get("agency", "-")],
        ["Media Type", campaign.get("mediaType", "-"), "City",    campaign.get("city", "-")],
        ["Start",      campaign.get("startDate", "-"), "End",     campaign.get("endDate", "-")],
        ["Budget",     f"₹ {campaign.get('budget', 0):,}", "Spent", f"₹ {campaign.get('spent', 0):,}"],
    ]
    meta_table = Table(meta_data, colWidths=[3.2 * cm, 5.5 * cm, 3.2 * cm, 5.5 * cm])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f8fafc")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR",  (0, 0), (0, -1), colors.HexColor("#64748b")),
        ("TEXTCOLOR",  (2, 0), (2, -1), colors.HexColor("#64748b")),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 18))

    total      = campaign.get("totalTasks", 0) or 0
    completed  = campaign.get("completed",  0) or 0
    flagged    = campaign.get("flagged",    0) or 0
    completion = (completed / total * 100) if total else 0

    kpis = [
        ["Total Tasks", "Completed", "Completion %", "Flagged"],
        [str(total),    str(completed), f"{completion:.1f}%", str(flagged)],
    ]
    kpi_table = Table(kpis, colWidths=[4.35 * cm] * 4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_color),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("FONTSIZE",   (0, 1), (-1, 1), 16),
        ("FONTNAME",   (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR",  (0, 1), (-1, 1), colors.HexColor("#0f172a")),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 18))

    story.append(Paragraph(
        "<b>Task Details</b>",
        ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=13),
    ))
    story.append(Spacer(1, 6))

    header = ["Task Code", "Unit", "City", "Executive", "Status", "Submitted"]
    rows = [header]
    for t in tasks[:200]:  # Full report — not capped at 35
        rows.append([
            t.get("taskCode", "-"),
            t.get("unitCode",  "-"),
            t.get("city",      "-"),
            (t.get("assignedTo", "-") or "-")[:22],
            (t.get("status",   "-") or "-").replace("_", " ").title(),
            (t.get("submittedAt") or "-")[:16],
        ])
    task_table = Table(
        rows,
        colWidths=[3.2 * cm, 1.8 * cm, 2.6 * cm, 4.2 * cm, 2.4 * cm, 3.2 * cm],
    )
    task_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.HexColor("#334155")),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(task_table)
    story.append(Spacer(1, 14))

    footer = (
        f"Generated by Moviq &nbsp;·&nbsp; "
        f"{datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}"
    )
    story.append(Paragraph(
        footer,
        ParagraphStyle(name="Foot", parent=styles["Normal"],
                       fontSize=8, textColor=colors.HexColor("#94a3b8"), alignment=1),
    ))
    doc.build(story)
    buf.seek(0)
    return buf.read()


def build_excel(campaign: Dict[str, Any], tasks: List[Dict[str, Any]]) -> bytes:
    """Return an Excel (.xlsx) campaign report as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    brand_fill  = PatternFill("solid", fgColor="DC2626")
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    thin   = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold       = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    ws["A1"] = "MOVIQ"
    ws["A1"].font = Font(bold=True, size=14, color="DC2626")
    ws["A2"] = "Campaign Report"
    ws["A2"].font = Font(bold=True, size=18)
    ws["A3"] = campaign.get("title", "")
    ws["A3"].font = Font(size=12, color="475569")

    meta = [
        ("Brand",       campaign.get("brand",       "-")),
        ("Agency",      campaign.get("agency",      "-")),
        ("Media Type",  campaign.get("mediaType",   "-")),
        ("City",        campaign.get("city",        "-")),
        ("Start Date",  campaign.get("startDate",   "-")),
        ("End Date",    campaign.get("endDate",     "-")),
        ("Budget (₹)",  campaign.get("budget",       0)),
        ("Spent (₹)",   campaign.get("spent",        0)),
        ("Total Tasks", campaign.get("totalTasks",   0)),
        ("Completed",   campaign.get("completed",    0)),
        ("Flagged",     campaign.get("flagged",      0)),
    ]
    row = 5
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=1).fill   = header_fill
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=val).border = border
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42

    ts = wb.create_sheet("Tasks")
    headers = [
        "Task Code", "Unit", "City", "Media Type",
        "Executive", "Status", "Submitted",
        "GPS Lat", "GPS Lng", "Photos", "Flag Reason",
    ]
    for i, h in enumerate(headers, 1):
        cell = ts.cell(row=1, column=i, value=h)
        cell.font      = white_bold
        cell.fill      = brand_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border

    widths = [16, 10, 14, 18, 20, 14, 20, 12, 12, 8, 30]
    for i, w in enumerate(widths, 1):
        ts.column_dimensions[chr(64 + i)].width = w

    for r, t in enumerate(tasks, start=2):
        vals = [
            t.get("taskCode", ""),   t.get("unitCode", ""),
            t.get("city", ""),       t.get("mediaType", ""),
            t.get("assignedTo", ""), t.get("status", ""),
            t.get("submittedAt", ""),t.get("lat", ""),
            t.get("lng", ""),        t.get("photos", 0),
            t.get("flagReason", "") or "",
        ]
        for i, v in enumerate(vals, 1):
            ts.cell(row=r, column=i, value=v).border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
