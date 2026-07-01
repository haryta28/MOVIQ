from fastapi import FastAPI, APIRouter, Depends, HTTPException
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
from pathlib import Path
from pydantic import BaseModel, EmailStr
from typing import List, Optional, Any, Dict
import uuid
from datetime import datetime, timedelta, timezone
import jwt
from passlib.context import CryptContext

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from seed_data import (
    AGENCIES, BRANDS, MEDIA_TYPES, CAMPAIGNS, TASKS, FRAUD_ALERTS,
    FIELD_EXECUTIVES, SUPERVISORS, MONTHLY_STATS, CITY_STATS, NOTIFICATIONS,
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ---------- Config ----------
JWT_SECRET = os.environ.get("JWT_SECRET", "moviq-super-secret-key-change-in-prod")
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = 24 * 7

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------- Mongo ----------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ---------- App ----------
app = FastAPI(title="Moviq API")
api = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)


# ---------- Models ----------
class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class AgencyCreate(BaseModel):
    name: str
    head: str = ""
    email: EmailStr
    phone: str = ""
    city: str = ""
    plan: str = "Growth"


class CampaignCreate(BaseModel):
    title: str
    brand: str
    brandId: Optional[str] = None
    mediaType: str
    city: str
    totalTasks: int = 0
    budget: int = 0
    startDate: str = ""
    endDate: str = ""


class MediaTypeCreate(BaseModel):
    label: str
    category: str = "Custom"


class TaskUpdate(BaseModel):
    status: Optional[str] = None
    flagReason: Optional[str] = None


class VehicleSubmissionCreate(BaseModel):
    vehicle: str
    driver_name: str
    driver_phone: str
    photos: List[Dict[str, str]] = []
    gps: Optional[Dict[str, float]] = None


# ---------- Helpers ----------
def _clean(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not doc:
        return doc
    doc.pop("_id", None)
    return doc


def _clean_many(docs):
    return [_clean(d) for d in docs]


def create_jwt(user: Dict[str, Any]) -> str:
    payload = {
        "sub": user["id"],
        "email": user["email"],
        "role": user["role"],
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(creds: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Dict[str, Any]:
    if not creds:
        raise HTTPException(status_code=401, detail="Missing token")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"id": payload["sub"]})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    user.pop("password_hash", None)
    return _clean(user)


async def require_admin(user: Dict = Depends(get_current_user)) -> Dict[str, Any]:
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user


async def create_notification(title: str, description: str, ntype: str = "info"):
    """Create a notification record. type: alert | success | info"""
    doc = {
        "id": f"n_{uuid.uuid4().hex[:10]}",
        "title": title,
        "description": description,
        "type": ntype,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "time": "just now",
    }
    await db.notifications.insert_one({**doc})
    return doc


def _relative_time(iso_str: Optional[str]) -> str:
    if not iso_str:
        return "just now"
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
    except Exception:
        return iso_str
    delta = datetime.now(timezone.utc) - dt
    secs = int(delta.total_seconds())
    if secs < 60:
        return "just now"
    if secs < 3600:
        return f"{secs // 60}m ago"
    if secs < 86400:
        return f"{secs // 3600}h ago"
    return f"{secs // 86400}d ago"


# ---------- Seeding ----------
async def _seed_if_empty(coll_name: str, docs: List[Dict[str, Any]]):
    coll = db[coll_name]
    if await coll.count_documents({}) == 0 and docs:
        await coll.insert_many([{**d} for d in docs])
        logger.info(f"Seeded {coll_name}: {len(docs)} docs")


async def seed_all():
    await _seed_if_empty("agencies", AGENCIES)
    await _seed_if_empty("brands", BRANDS)
    await _seed_if_empty("media_types", MEDIA_TYPES)
    await _seed_if_empty("campaigns", CAMPAIGNS)
    await _seed_if_empty("tasks", TASKS)
    await _seed_if_empty("fraud_alerts", FRAUD_ALERTS)
    await _seed_if_empty("field_executives", FIELD_EXECUTIVES)
    await _seed_if_empty("supervisors", SUPERVISORS)
    await _seed_if_empty("monthly_stats", MONTHLY_STATS)
    await _seed_if_empty("city_stats", CITY_STATS)
    await _seed_if_empty("notifications", NOTIFICATIONS)

    # Backfill createdAt on notifications that lack it (from old seed)
    now = datetime.now(timezone.utc)
    async for n in db.notifications.find({"createdAt": {"$exists": False}}):
        # Guess timestamp based on the mock 'time' string; fallback to now - index*1h
        await db.notifications.update_one(
            {"id": n["id"]},
            {"$set": {"createdAt": (now - timedelta(hours=1)).isoformat()}}
        )

    if await db.users.count_documents({}) == 0:
        seed_users = [
            {"id": "u1", "name": "Deepak Bansal", "email": "admin@moviq.in", "role": "admin", "avatar": "DB",
             "password_hash": pwd_ctx.hash("demo1234")},
            {"id": "u2", "name": "Saurav Mehta", "email": "saurav@brightads.in", "role": "agency", "avatar": "SM",
             "agencyId": "a1", "agencyName": "BrightAds Media",
             "password_hash": pwd_ctx.hash("demo1234")},
        ]
        await db.users.insert_many(seed_users)
        logger.info(f"Seeded users: {len(seed_users)}")


@app.on_event("startup")
async def on_startup():
    await seed_all()


@app.on_event("shutdown")
async def on_shutdown():
    client.close()


# ---------- Routes ----------
@api.get("/")
async def root():
    return {"message": "Moviq API", "status": "ok"}


# ---- Auth ----
@api.post("/auth/login")
async def auth_login(body: LoginRequest):
    user = await db.users.find_one({"email": body.email.lower()})
    if not user or not pwd_ctx.verify(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    user.pop("password_hash", None)
    user = _clean(user)
    return {"token": create_jwt(user), "user": user}


@api.get("/auth/me")
async def auth_me(user: Dict = Depends(get_current_user)):
    return user


# ---- Agencies ----
@api.get("/agencies")
async def list_agencies(_: Dict = Depends(get_current_user)):
    return _clean_many(await db.agencies.find().to_list(1000))


@api.post("/agencies")
async def create_agency(body: AgencyCreate, _: Dict = Depends(require_admin)):
    new = {
        "id": f"a{uuid.uuid4().hex[:8]}",
        **body.dict(),
        "campaigns": 0, "activeUsers": 1, "status": "trial", "revenue": 0,
        "joinedAt": datetime.now(timezone.utc).date().isoformat(),
    }
    await db.agencies.insert_one(new)
    await create_notification(
        "Agency onboarded",
        f"{new['name']} joined on the {new.get('plan','Growth')} plan",
        "info",
    )
    return _clean(new)


@api.get("/agencies/{aid}")
async def get_agency(aid: str, _: Dict = Depends(get_current_user)):
    doc = await db.agencies.find_one({"id": aid})
    if not doc:
        raise HTTPException(status_code=404, detail="Agency not found")
    return _clean(doc)


# ---- Brands ----
@api.get("/brands")
async def list_brands(_: Dict = Depends(get_current_user)):
    return _clean_many(await db.brands.find().to_list(1000))


# ---- Campaigns ----
@api.get("/campaigns")
async def list_campaigns(agency_id: Optional[str] = None, user: Dict = Depends(get_current_user)):
    q: Dict[str, Any] = {}
    if agency_id:
        q["agencyId"] = agency_id
    elif user["role"] == "agency":
        q["agencyId"] = user.get("agencyId")
    return _clean_many(await db.campaigns.find(q).to_list(1000))


@api.get("/campaigns/{cid}")
async def get_campaign(cid: str, user: Dict = Depends(get_current_user)):
    doc = await db.campaigns.find_one({"id": cid})
    if not doc:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if user["role"] == "agency" and doc.get("agencyId") != user.get("agencyId"):
        raise HTTPException(status_code=403, detail="Not your campaign")

    tasks = await db.tasks.find({"campaignId": cid}).to_list(2000)
    field = await db.field_executives.find({"agencyId": doc.get("agencyId")}).to_list(200)
    fraud = await db.fraud_alerts.find({"taskCode": {"$in": [t.get("taskCode") for t in tasks]}}).to_list(200)

    # Simple derived activity feed
    activity = []
    for t in tasks[:12]:
        if t.get("submittedAt"):
            activity.append({
                "id": t["id"],
                "kind": "submission",
                "text": f"{t.get('assignedTo','')} submitted proof for {t.get('unitCode','')}",
                "time": t.get("submittedAt"),
                "status": t.get("status"),
            })
    for f in fraud:
        activity.append({
            "id": f["id"],
            "kind": "fraud",
            "text": f"Fraud alert: {f.get('type','')} on {f.get('taskCode','')}",
            "time": f.get("detectedAt"),
            "status": "flagged",
        })

    return {
        "campaign": _clean(doc),
        "tasks": _clean_many(tasks),
        "team": _clean_many(field),
        "activity": activity,
        "stats": {
            "byStatus": {
                s: sum(1 for t in tasks if t.get("status") == s)
                for s in ["pending", "in_progress", "submitted", "approved", "flagged"]
            },
            "byCity": {},
        },
    }


@api.post("/campaigns")
async def create_campaign(body: CampaignCreate, user: Dict = Depends(get_current_user)):
    if user["role"] != "agency":
        raise HTTPException(status_code=403, detail="Only agency users can create campaigns")
    new = {
        "id": f"c{uuid.uuid4().hex[:8]}",
        **body.dict(),
        "agency": user.get("agencyName", ""),
        "agencyId": user.get("agencyId"),
        "completed": 0, "flagged": 0, "status": "ongoing", "spent": 0,
    }
    await db.campaigns.insert_one(new)
    await create_notification(
        "Campaign launched",
        f"{new['title']} for {new.get('brand','')} is now live in {new.get('city','')}",
        "success",
    )
    return _clean(new)


# ---- Tasks ----
@api.get("/tasks")
async def list_tasks(agency_id: Optional[str] = None, status_: Optional[str] = None, city: Optional[str] = None,
                     user: Dict = Depends(get_current_user)):
    q: Dict[str, Any] = {}
    if agency_id:
        q["agencyId"] = agency_id
    elif user["role"] == "agency":
        q["agencyId"] = user.get("agencyId")
    if status_:
        q["status"] = status_
    if city:
        q["city"] = city
    return _clean_many(await db.tasks.find(q).to_list(2000))


@api.patch("/tasks/{tid}")
async def update_task(tid: str, body: TaskUpdate, user: Dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": tid})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if user["role"] == "agency" and task.get("agencyId") != user.get("agencyId"):
        raise HTTPException(status_code=403, detail="Not your task")
    update: Dict[str, Any] = {}
    if body.status is not None:
        update["status"] = body.status
        if body.status == "flagged" and body.flagReason:
            update["flagReason"] = body.flagReason
        elif body.status != "flagged":
            update["flagReason"] = None
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.tasks.update_one({"id": tid}, {"$set": update})
    return _clean(await db.tasks.find_one({"id": tid}))


# ---- Users ----
@api.get("/users")
async def list_users(role: Optional[str] = None, _: Dict = Depends(get_current_user)):
    if role == "field":
        return _clean_many(await db.field_executives.find().to_list(1000))
    if role == "supervisor":
        return _clean_many(await db.supervisors.find().to_list(1000))
    if role == "agency":
        agencies = await db.agencies.find().to_list(1000)
        return [{"id": a["id"], "name": a["head"], "email": a["email"], "agency": a["name"], "status": a["status"]} for a in agencies]
    if role == "admin":
        return [
            {"id": "ad1", "name": "Deepak Bansal", "email": "admin@moviq.in", "role": "Super Admin", "status": "active"},
            {"id": "ad2", "name": "Anjali Sharma", "email": "anjali@moviq.in", "role": "Operations", "status": "active"},
            {"id": "ad3", "name": "Rohan Iyer", "email": "rohan@moviq.in", "role": "Support", "status": "active"},
        ]
    return _clean_many(await db.users.find({}, {"password_hash": 0}).to_list(1000))


# ---- Fraud alerts ----
@api.get("/fraud-alerts")
async def list_fraud(_: Dict = Depends(get_current_user)):
    return _clean_many(await db.fraud_alerts.find().to_list(1000))


@api.post("/fraud-alerts/{fid}/resolve")
async def resolve_fraud(fid: str, _: Dict = Depends(get_current_user)):
    alert = await db.fraud_alerts.find_one({"id": fid})
    if not alert:
        raise HTTPException(status_code=404, detail="Alert not found")
    await db.fraud_alerts.delete_one({"id": fid})
    await create_notification(
        "Fraud alert resolved",
        f"{alert.get('type','')} on {alert.get('taskCode','')} marked as reviewed",
        "success",
    )
    return {"ok": True}


# ---- Media types ----
@api.get("/media-types")
async def list_media_types(_: Dict = Depends(get_current_user)):
    return _clean_many(await db.media_types.find().to_list(1000))


@api.post("/media-types")
async def create_media_type(body: MediaTypeCreate, _: Dict = Depends(require_admin)):
    key = body.label.lower().replace(" ", "_")
    if await db.media_types.find_one({"key": key}):
        raise HTTPException(status_code=409, detail="Media type already exists")
    doc = {"key": key, "label": body.label, "category": body.category}
    await db.media_types.insert_one(doc)
    return _clean(doc)


@api.delete("/media-types/{mkey}")
async def delete_media_type(mkey: str, _: Dict = Depends(require_admin)):
    res = await db.media_types.delete_one({"key": mkey})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Media type not found")
    return {"ok": True}


# ---- Vehicle Submissions (WhatsApp bot) ----
@api.post("/vehicle-submissions")
async def create_vehicle_submission(body: VehicleSubmissionCreate):
    """Public endpoint used by the WhatsApp bot simulator. No auth required."""
    doc = {
        "id": f"vs_{uuid.uuid4().hex[:10]}",
        "vehicle": body.vehicle.upper(),
        "driverName": body.driver_name,
        "driverPhone": body.driver_phone,
        "photos": body.photos,
        "gps": body.gps or {"lat": 12.9784, "lng": 77.5946},
        "submittedAt": datetime.now(timezone.utc).isoformat(),
        "status": "submitted",
        "fraudCheck": "passed",
    }
    await db.vehicle_submissions.insert_one({**doc})
    await create_notification(
        "New vehicle proof",
        f"{doc['vehicle']} registered by {doc['driverName']}",
        "info",
    )
    return _clean(doc)


@api.get("/vehicle-submissions")
async def list_vehicle_submissions(_: Dict = Depends(get_current_user)):
    docs = await db.vehicle_submissions.find().sort("submittedAt", -1).to_list(500)
    return _clean_many(docs)


# ---- Analytics ----
@api.get("/analytics/overview")
async def analytics_overview(user: Dict = Depends(get_current_user)):
    monthly = _clean_many(await db.monthly_stats.find().to_list(100))
    cities = _clean_many(await db.city_stats.find().to_list(100))

    agencies = await db.agencies.find().to_list(1000)
    campaigns = await db.campaigns.find().to_list(1000)

    if user["role"] == "agency":
        campaigns = [c for c in campaigns if c.get("agencyId") == user.get("agencyId")]

    total_revenue = sum(a.get("revenue", 0) for a in agencies)
    completed = sum(c.get("completed", 0) for c in campaigns)
    total = sum(c.get("totalTasks", 0) for c in campaigns)
    active_agencies = sum(1 for a in agencies if a.get("status") == "active")
    live_campaigns = sum(1 for c in campaigns if c.get("status") == "ongoing")

    return {
        "monthlyStats": monthly,
        "cityStats": cities,
        "kpis": {
            "activeAgencies": active_agencies,
            "liveCampaigns": live_campaigns,
            "tasksExecuted": completed,
            "totalTasks": total,
            "totalRevenue": total_revenue,
        },
    }


# ---- Notifications ----
@api.get("/notifications")
async def list_notifications(_: Dict = Depends(get_current_user)):
    docs = await db.notifications.find().sort("createdAt", -1).to_list(50)
    out = _clean_many(docs)
    for n in out:
        if n.get("createdAt"):
            n["time"] = _relative_time(n.get("createdAt"))
    return out


# ---- Reports (PDF / Excel) ----
def _build_pdf(campaign: Dict[str, Any], tasks: List[Dict[str, Any]]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.4 * cm, rightMargin=1.4 * cm,
                            topMargin=1.4 * cm, bottomMargin=1.4 * cm)
    styles = getSampleStyleSheet()
    story = []
    brand_color = colors.HexColor("#DC2626")

    styles.add(ParagraphStyle(name="Brand", parent=styles["Normal"], fontSize=10, textColor=brand_color, fontName="Helvetica-Bold"))
    story.append(Paragraph("MOVIQ &nbsp;·&nbsp; Intelligence in Motion", styles["Brand"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"<b>Campaign Report</b>", ParagraphStyle(name="H1", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#0f172a"))))
    story.append(Paragraph(campaign.get("title", ""), ParagraphStyle(name="Sub", parent=styles["Normal"], fontSize=13, textColor=colors.HexColor("#334155"))))
    story.append(Spacer(1, 14))

    meta_data = [
        ["Brand", campaign.get("brand", "-"), "Agency", campaign.get("agency", "-")],
        ["Media Type", campaign.get("mediaType", "-"), "City", campaign.get("city", "-")],
        ["Start", campaign.get("startDate", "-"), "End", campaign.get("endDate", "-")],
        ["Budget", f"₹ {campaign.get('budget', 0):,}", "Spent", f"₹ {campaign.get('spent', 0):,}"],
    ]
    meta_table = Table(meta_data, colWidths=[3.2 * cm, 5.5 * cm, 3.2 * cm, 5.5 * cm])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f8fafc")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#64748b")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#64748b")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 18))

    total = campaign.get("totalTasks", 0) or 0
    completed = campaign.get("completed", 0) or 0
    flagged = campaign.get("flagged", 0) or 0
    completion = (completed / total * 100) if total else 0

    kpis = [["Total Tasks", "Completed", "Completion %", "Flagged"],
            [str(total), str(completed), f"{completion:.1f}%", str(flagged)]]
    kpi_table = Table(kpis, colWidths=[4.35 * cm] * 4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 16),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#0f172a")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 18))

    story.append(Paragraph("<b>Task Details</b>", ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=13)))
    story.append(Spacer(1, 6))

    header = ["Task Code", "Unit", "City", "Executive", "Status", "Submitted"]
    rows = [header]
    for t in tasks[:35]:
        rows.append([
            t.get("taskCode", "-"),
            t.get("unitCode", "-"),
            t.get("city", "-"),
            (t.get("assignedTo", "-") or "-")[:22],
            (t.get("status", "-") or "-").replace("_", " ").title(),
            (t.get("submittedAt") or "-")[:16],
        ])
    task_table = Table(rows, colWidths=[3.2 * cm, 1.8 * cm, 2.6 * cm, 4.2 * cm, 2.4 * cm, 3.2 * cm])
    task_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#334155")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(task_table)
    story.append(Spacer(1, 14))

    footer = f"Generated by Moviq &nbsp;·&nbsp; {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}"
    story.append(Paragraph(footer, ParagraphStyle(name="Foot", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#94a3b8"), alignment=1)))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _build_excel(campaign: Dict[str, Any], tasks: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    brand_fill = PatternFill("solid", fgColor="DC2626")
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    ws["A1"] = "MOVIQ"
    ws["A1"].font = Font(bold=True, size=14, color="DC2626")
    ws["A2"] = "Campaign Report"
    ws["A2"].font = Font(bold=True, size=18)
    ws["A3"] = campaign.get("title", "")
    ws["A3"].font = Font(size=12, color="475569")

    meta = [
        ("Brand", campaign.get("brand", "-")),
        ("Agency", campaign.get("agency", "-")),
        ("Media Type", campaign.get("mediaType", "-")),
        ("City", campaign.get("city", "-")),
        ("Start Date", campaign.get("startDate", "-")),
        ("End Date", campaign.get("endDate", "-")),
        ("Budget (₹)", campaign.get("budget", 0)),
        ("Spent (₹)", campaign.get("spent", 0)),
        ("Total Tasks", campaign.get("totalTasks", 0)),
        ("Completed", campaign.get("completed", 0)),
        ("Flagged", campaign.get("flagged", 0)),
    ]
    row = 5
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=val).border = border
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42

    # Tasks sheet
    ts = wb.create_sheet("Tasks")
    headers = ["Task Code", "Unit", "City", "Media Type", "Executive", "Status", "Submitted", "GPS Lat", "GPS Lng", "Photos", "Flag Reason"]
    for i, h in enumerate(headers, 1):
        cell = ts.cell(row=1, column=i, value=h)
        cell.font = white_bold
        cell.fill = brand_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    widths = [16, 10, 14, 18, 20, 14, 20, 12, 12, 8, 30]
    for i, w in enumerate(widths, 1):
        ts.column_dimensions[chr(64 + i)].width = w

    for r, t in enumerate(tasks, start=2):
        vals = [
            t.get("taskCode", ""), t.get("unitCode", ""), t.get("city", ""),
            t.get("mediaType", ""), t.get("assignedTo", ""), t.get("status", ""),
            t.get("submittedAt", ""), t.get("lat", ""), t.get("lng", ""),
            t.get("photos", 0), t.get("flagReason", "") or "",
        ]
        for i, v in enumerate(vals, 1):
            cell = ts.cell(row=r, column=i, value=v)
            cell.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _campaign_with_tasks(cid: str, user: Dict[str, Any]):
    doc = await db.campaigns.find_one({"id": cid})
    if not doc:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if user["role"] == "agency" and doc.get("agencyId") != user.get("agencyId"):
        raise HTTPException(status_code=403, detail="Not your campaign")
    tasks = await db.tasks.find({"campaignId": cid}).to_list(2000)
    return _clean(doc), _clean_many(tasks)


@api.get("/campaigns/{cid}/report/pdf")
async def campaign_report_pdf(cid: str, user: Dict = Depends(get_current_user)):
    campaign, tasks = await _campaign_with_tasks(cid, user)
    pdf_bytes = _build_pdf(campaign, tasks)
    filename = f"moviq-report-{cid}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api.get("/campaigns/{cid}/report/excel")
async def campaign_report_excel(cid: str, user: Dict = Depends(get_current_user)):
    campaign, tasks = await _campaign_with_tasks(cid, user)
    xlsx_bytes = _build_excel(campaign, tasks)
    filename = f"moviq-report-{cid}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Include & CORS ----------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
